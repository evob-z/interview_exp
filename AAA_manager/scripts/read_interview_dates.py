"""
读取简历投递 Excel，输出面试日期映射表。

用法：
    python read_interview_dates.py
    python read_interview_dates.py --xlsx "C:/path/to/simli.xlsx"

产物：
    1. stdout 打印人工可读的表格
    2. 同目录下写出 .interview_dates.json 供后续脚本/Skill 复用
"""

from __future__ import annotations

import argparse
import json
import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "缺少依赖 openpyxl，请执行: pip install openpyxl"
    ) from exc


DEFAULT_XLSX = os.getenv("COMPANY_EXCEL_PATH", "")

# 默认输出到项目仓库根目录
_REPO = os.getenv("INTERVIEW_REPO_PATH", str(Path(__file__).resolve().parent.parent.parent))
DEFAULT_JSON_OUT = Path(_REPO) / ".interview_dates.json"

# 列名候选（越靠前优先级越高）
COL_ALIASES: dict[str, list[str]] = {
    "company": ["公司", "公司名称", "企业", "单位"],
    "category": ["类型", "公司类型", "规模", "性质"],
    "date": ["面试日期", "日期", "时间", "面试时间"],
    "round": ["轮次", "场次", "面试轮次", "哪场面试", "面试"],
}


def normalize_date(value: Any) -> str | None:
    """把各种日期表达归一化为 YYMMDD。"""
    if value is None or value == "":
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%y%m%d")
    text = str(value).strip()
    # 尝试常见格式
    patterns = [
        "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d",
        "%y-%m-%d", "%y/%m/%d", "%y.%m.%d",
        "%Y年%m月%d日", "%m月%d日",
    ]
    for p in patterns:
        try:
            return datetime.strptime(text, p).strftime("%y%m%d")
        except ValueError:
            continue
    # 纯数字 260130 形式
    if re.fullmatch(r"\d{6}", text):
        return text
    if re.fullmatch(r"\d{8}", text):
        return text[2:]
    return None


def detect_columns(header_row: list[Any]) -> dict[str, int]:
    """根据表头匹配列索引。"""
    mapping: dict[str, int] = {}
    header_str = [str(h).strip() if h is not None else "" for h in header_row]
    for key, aliases in COL_ALIASES.items():
        for alias in aliases:
            for idx, cell in enumerate(header_str):
                if alias in cell:
                    mapping[key] = idx
                    break
            if key in mapping:
                break
    return mapping


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="Excel 路径")
    parser.add_argument("--out", default=str(DEFAULT_JSON_OUT), help="JSON 输出路径")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise SystemExit(f"找不到 Excel: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    records: list[dict[str, Any]] = []

    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        # 探测表头所在行（前 3 行里找）
        header_row = None
        header_idx = 0
        for i, row in enumerate(rows[:3]):
            mapping_try = detect_columns(list(row))
            if "company" in mapping_try and "date" in mapping_try:
                header_row = list(row)
                header_idx = i
                break
        if header_row is None:
            continue
        mapping = detect_columns(header_row)

        for row in rows[header_idx + 1:]:
            if not any(c is not None and str(c).strip() for c in row):
                continue
            company = row[mapping["company"]] if "company" in mapping else None
            if not company:
                continue
            category = row[mapping["category"]] if "category" in mapping else ""
            raw_date = row[mapping["date"]] if "date" in mapping else None
            round_ = row[mapping["round"]] if "round" in mapping else ""
            yymmdd = normalize_date(raw_date)
            records.append({
                "sheet": sheet.title,
                "company": str(company).strip(),
                "category": str(category).strip() if category else "",
                "date_raw": str(raw_date) if raw_date is not None else "",
                "date": yymmdd or "",
                "round": str(round_).strip() if round_ else "",
            })

    # 打印表格
    print(f"{'公司':<20} {'类型':<8} {'YYMMDD':<8} {'轮次':<10} 原始日期")
    print("-" * 70)
    for r in records:
        print(
            f"{r['company']:<20} {r['category']:<8} {r['date']:<8} "
            f"{r['round']:<10} {r['date_raw']}"
        )

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(records, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"\n共 {len(records)} 条记录，已写入 {out_path}")


if __name__ == "__main__":
    main()
